from __future__ import annotations

import csv
import io
import logging
import re
import uuid
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert as postgresql_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.orm import Session

from .models import (
    AdminSessionLock,
    Event,
    FaceParticipantMatch,
    FaceSearchImage,
    FaceSearchJob,
    FaceSearchResult,
    FaceSearchSession,
    Participant,
    ParticipantFaceEmbedding,
    ParticipantFaceImage,
    ParticipantFaceJob,
    PhotoParticipantMatch,
)
from .config import settings
from .maintenance import enqueue_object_deletion, process_object_deletions
from .participant_lookup import normalize_bib_lookup, normalize_name_lookup


logger = logging.getLogger(__name__)


EVENT_STATUSES = ("draft", "published")
ADMIN_LOCK_ID = 1
ADMIN_SESSION_TTL = timedelta(minutes=30)

HEADER_ALIASES = {
    "bib_number": {
        "bib",
        "bib_number",
        "bib_no",
        "bib_num",
        "bib_no.",
        "bib#",
        "bib_number.",
        "race_number",
        "runner_number",
        "number",
    },
    "full_name": {
        "full_name",
        "name",
        "participant_name",
        "runner_name",
        "athlete_name",
    },
    "first_name": {
        "first_name",
        "firstname",
        "given_name",
    },
    "last_name": {
        "last_name",
        "lastname",
        "surname",
        "family_name",
    },
}


@dataclass
class EventListItem:
    event: Event
    participant_count: int


@dataclass
class ParticipantUploadResult:
    processed_rows: int
    inserted_rows: int
    updated_rows: int
    skipped_rows: int


def slugify(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return normalized or "event"


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    return text


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def list_events(session: Session) -> list[EventListItem]:
    stmt = (
        select(Event, func.count(Participant.id))
        .outerjoin(Participant, Participant.event_id == Event.id)
        .group_by(Event.id)
        .order_by(Event.created_at.desc())
    )
    return [EventListItem(event=row[0], participant_count=row[1]) for row in session.execute(stmt).all()]


def list_participants(session: Session, *, event_id: uuid.UUID) -> list[Participant]:
    return (
        session.query(Participant)
        .filter(Participant.event_id == event_id)
        .order_by(Participant.bib_number.asc(), Participant.created_at.asc())
        .all()
    )


def get_event(session: Session, event_id: str) -> Event | None:
    try:
        parsed_id = uuid.UUID(str(event_id))
    except ValueError:
        return None
    return session.get(Event, parsed_id)


def get_participant(session: Session, participant_id: str) -> Participant | None:
    try:
        parsed_id = uuid.UUID(str(participant_id))
    except ValueError:
        return None
    return session.get(Participant, parsed_id)


def slug_exists(session: Session, slug: str, *, exclude_event_id: str | None = None) -> bool:
    stmt = select(Event.id).where(Event.slug == slug)
    if exclude_event_id:
        try:
            stmt = stmt.where(Event.id != uuid.UUID(str(exclude_event_id)))
        except ValueError:
            pass
    return session.execute(stmt).scalar_one_or_none() is not None


def generate_unique_slug(session: Session, *, name: str, exclude_event_id: str | None = None) -> str:
    base_slug = slugify(name)
    candidate = base_slug
    suffix = 2
    while slug_exists(session, candidate, exclude_event_id=exclude_event_id):
        candidate = f"{base_slug}-{suffix}"
        suffix += 1
    return candidate


def create_event(
    session: Session,
    *,
    name: str,
    slug: str,
    event_date: date | None,
    location: str | None,
    status: str,
) -> Event:
    event = Event(
        name=name,
        slug=slug,
        event_date=event_date,
        location=location,
        status=status,
    )
    session.add(event)
    session.commit()
    session.refresh(event)
    return event


def update_event(
    session: Session,
    *,
    event: Event,
    name: str,
    slug: str,
    event_date: date | None,
    location: str | None,
    status: str,
) -> Event:
    event.name = name
    event.slug = slug
    event.event_date = event_date
    event.location = location
    event.status = status
    session.commit()
    session.refresh(event)
    return event


def delete_event(session: Session, *, event: Event) -> None:
    object_keys = list(
        session.execute(select(FaceSearchImage.object_key).where(FaceSearchImage.event_id == event.id)).scalars()
    )
    object_keys.extend(
        session.execute(select(ParticipantFaceImage.object_key).where(ParticipantFaceImage.event_id == event.id)).scalars()
    )
    for object_key in object_keys:
        enqueue_object_deletion(session, object_key)
    session.delete(event)
    session.commit()
    process_object_deletions(session, limit=min(len(object_keys), settings.deletion_retry_batch_size))


def parse_participant_file(file_name: str, content: bytes) -> list[dict[str, str]]:
    if not content:
        raise ValueError("Participant upload is empty.")
    if len(content) > settings.max_participant_upload_bytes:
        raise ValueError("Participant upload exceeds the configured size limit.")
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix == "csv":
        return _parse_csv(content)
    if suffix in {"xlsx", "xlsm"}:
        return _parse_excel(content)
    raise ValueError("Upload a CSV or XLSX file.")


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValueError("CSV must be UTF-8 encoded.") from None
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is missing a header row.")
    if len(reader.fieldnames) > settings.max_participant_columns:
        raise ValueError("Participant file has too many columns.")
    rows: list[dict[str, str]] = []
    for index, row in enumerate(reader, start=1):
        if index > settings.max_participant_rows:
            raise ValueError("Participant file has too many rows.")
        rows.append(_normalize_row(row))
    return rows


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > 2_000:
                raise ValueError("Spreadsheet archive contains too many files.")
            expanded_size = sum(entry.file_size for entry in entries)
            compressed_size = max(1, sum(entry.compress_size for entry in entries))
            if expanded_size > settings.max_spreadsheet_uncompressed_bytes or expanded_size / compressed_size > 100:
                raise ValueError("Spreadsheet archive expands beyond the configured safety limit.")
    except zipfile.BadZipFile:
        raise ValueError("Spreadsheet is not a valid XLSX file.") from None

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            raise ValueError("Spreadsheet is empty.")
        if len(first_row) > settings.max_participant_columns:
            raise ValueError("Participant file has too many columns.")
        headers = [normalize_header(cell) for cell in first_row]
        if not any(headers):
            raise ValueError("Spreadsheet is missing a header row.")
        normalized_rows: list[dict[str, str]] = []
        for row_index, values in enumerate(rows, start=1):
            if row_index > settings.max_participant_rows:
                raise ValueError("Participant file has too many rows.")
            if len(values) > settings.max_participant_columns:
                raise ValueError("Participant file has too many columns.")
            row = {headers[index]: "" if value is None else str(value).strip() for index, value in enumerate(values)}
            normalized_rows.append(_normalize_row(row))
        return normalized_rows
    except (OSError, KeyError, ValueError) as exc:
        if isinstance(exc, ValueError):
            raise
        raise ValueError("Spreadsheet could not be parsed safely.") from None
    finally:
        if "workbook" in locals():
            workbook.close()


def _normalize_row(raw_row: dict[object, object]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in raw_row.items():
        text = str(value or "").strip()
        if len(text) > settings.max_participant_cell_chars:
            raise ValueError("A participant cell exceeds the configured length limit.")
        normalized[normalize_header(key)] = text

    first_name = _read_value(normalized, "first_name")
    last_name = _read_value(normalized, "last_name")
    full_name = _read_value(normalized, "full_name")
    if not full_name and (first_name or last_name):
        full_name = f"{first_name} {last_name}".strip()

    return {
        "bib_number": _read_value(normalized, "bib_number"),
        "full_name": full_name,
    }


def _read_value(row: dict[str, str], canonical_name: str) -> str:
    for header in HEADER_ALIASES[canonical_name]:
        if header in row and row[header]:
            return row[header]
    return ""


def upsert_participants(
    session: Session,
    *,
    event: Event,
    rows: Iterable[dict[str, str]],
) -> ParticipantUploadResult:
    cleaned_rows: list[dict[str, str]] = []
    skipped_rows = 0
    for row in rows:
        bib_number = row.get("bib_number", "").strip()
        full_name = row.get("full_name", "").strip()
        if not bib_number or not full_name:
            skipped_rows += 1
            continue
        if len(bib_number) > 64 or len(full_name) > 255:
            skipped_rows += 1
            continue
        bib_lookup = normalize_bib_lookup(bib_number)
        if not bib_lookup:
            skipped_rows += 1
            continue
        cleaned_rows.append(
            {
                "bib_number": bib_number,
                "full_name": full_name,
                "bib_lookup": bib_lookup,
                "name_lookup": normalize_name_lookup(full_name),
            }
        )

    if not cleaned_rows:
        return ParticipantUploadResult(processed_rows=0, inserted_rows=0, updated_rows=0, skipped_rows=skipped_rows)

    existing = {
        participant.bib_lookup: participant
        for participant in session.execute(
            select(Participant).where(
                Participant.event_id == event.id,
                Participant.bib_lookup.in_([row["bib_lookup"] for row in cleaned_rows]),
            )
        ).scalars()
    }

    inserted_rows = 0
    updated_rows = 0
    deduped_by_bib = {row["bib_lookup"]: row for row in cleaned_rows}

    for bib_lookup, row in deduped_by_bib.items():
        participant = existing.get(bib_lookup)
        if participant is None:
            session.add(
                Participant(
                    event_id=event.id,
                    bib_number=row["bib_number"],
                    full_name=row["full_name"],
                    bib_lookup=bib_lookup,
                    name_lookup=row["name_lookup"],
                )
            )
            inserted_rows += 1
            continue
        if participant.full_name != row["full_name"] or participant.bib_number != row["bib_number"]:
            participant.bib_number = row["bib_number"]
            participant.full_name = row["full_name"]
            participant.bib_lookup = bib_lookup
            participant.name_lookup = row["name_lookup"]
            updated_rows += 1

    session.commit()
    duplicate_rows = len(cleaned_rows) - len(deduped_by_bib)
    return ParticipantUploadResult(
        processed_rows=len(deduped_by_bib),
        inserted_rows=inserted_rows,
        updated_rows=updated_rows,
        skipped_rows=skipped_rows + duplicate_rows,
    )


def add_participant(session: Session, *, event: Event, bib_number: str, full_name: str) -> Participant:
    normalized_bib = bib_number.strip()
    normalized_name = full_name.strip()
    bib_lookup = normalize_bib_lookup(normalized_bib)
    if not bib_lookup or not normalized_name:
        raise ValueError("Bib number and full name are required.")

    existing = (
        session.query(Participant)
        .filter(Participant.event_id == event.id, Participant.bib_lookup == bib_lookup)
        .one_or_none()
    )
    if existing is not None:
        raise ValueError("That bib number already exists for this event.")

    participant = Participant(
        event_id=event.id,
        bib_number=normalized_bib,
        full_name=normalized_name,
        bib_lookup=bib_lookup,
        name_lookup=normalize_name_lookup(normalized_name),
    )
    session.add(participant)
    session.commit()
    session.refresh(participant)
    return participant


def update_participant(
    session: Session,
    *,
    participant: Participant,
    bib_number: str,
    full_name: str,
) -> Participant:
    normalized_bib = bib_number.strip()
    normalized_name = full_name.strip()
    bib_lookup = normalize_bib_lookup(normalized_bib)
    if not bib_lookup or not normalized_name:
        raise ValueError("Bib number and full name are required.")

    existing = (
        session.query(Participant)
        .filter(
            Participant.event_id == participant.event_id,
            Participant.bib_lookup == bib_lookup,
            Participant.id != participant.id,
        )
        .one_or_none()
    )
    if existing is not None:
        raise ValueError("That bib number already exists for this event.")

    participant.bib_number = normalized_bib
    participant.full_name = normalized_name
    participant.bib_lookup = bib_lookup
    participant.name_lookup = normalize_name_lookup(normalized_name)
    session.commit()
    session.refresh(participant)
    return participant


def delete_participant(session: Session, *, participant: Participant) -> None:
    cleanup_participant_identity_data(session, participant_ids=[participant.id])
    session.delete(participant)
    session.commit()
    process_object_deletions(session)


def delete_all_participants(session: Session, *, event: Event) -> int:
    participant_ids = list(
        session.execute(
            select(Participant.id).where(Participant.event_id == event.id)
        ).scalars()
    )
    cleanup_participant_identity_data(session, participant_ids=participant_ids)
    count = (
        session.query(Participant)
        .filter(Participant.event_id == event.id)
        .delete(synchronize_session=False)
    )
    session.commit()
    process_object_deletions(session)
    return count


def cleanup_participant_identity_data(session: Session, *, participant_ids: Iterable[uuid.UUID]) -> None:
    ids = list(participant_ids)
    if not ids:
        return

    search_session_ids = list(
        session.execute(
            select(FaceSearchSession.id).where(FaceSearchSession.participant_id.in_(ids))
        ).scalars()
    )
    search_image_ids = []
    if search_session_ids:
        search_object_keys = list(
            session.execute(
                select(FaceSearchImage.object_key).where(FaceSearchImage.search_session_id.in_(search_session_ids))
            ).scalars()
        )
        for object_key in search_object_keys:
            enqueue_object_deletion(session, object_key)
        search_image_ids = list(
            session.execute(
                select(FaceSearchImage.id).where(FaceSearchImage.search_session_id.in_(search_session_ids))
            ).scalars()
        )
        session.query(FaceSearchResult).filter(FaceSearchResult.search_session_id.in_(search_session_ids)).delete(synchronize_session=False)
        if search_image_ids:
            session.query(FaceSearchJob).filter(FaceSearchJob.search_image_id.in_(search_image_ids)).delete(synchronize_session=False)
        session.query(FaceSearchJob).filter(FaceSearchJob.search_session_id.in_(search_session_ids)).delete(synchronize_session=False)
        session.query(FaceSearchImage).filter(FaceSearchImage.search_session_id.in_(search_session_ids)).delete(synchronize_session=False)
        session.query(FaceSearchSession).filter(FaceSearchSession.id.in_(search_session_ids)).delete(synchronize_session=False)

    face_images = list(
        session.execute(
            select(ParticipantFaceImage).where(ParticipantFaceImage.participant_id.in_(ids))
        ).scalars()
    )
    face_image_ids = [image.id for image in face_images]
    for image in face_images:
        enqueue_object_deletion(session, image.object_key)
    session.query(FaceParticipantMatch).filter(FaceParticipantMatch.participant_id.in_(ids)).delete(synchronize_session=False)
    session.query(PhotoParticipantMatch).filter(PhotoParticipantMatch.participant_id.in_(ids)).delete(synchronize_session=False)
    session.query(ParticipantFaceJob).filter(ParticipantFaceJob.participant_id.in_(ids)).delete(synchronize_session=False)
    session.query(ParticipantFaceEmbedding).filter(ParticipantFaceEmbedding.participant_id.in_(ids)).delete(synchronize_session=False)
    if face_image_ids:
        session.query(ParticipantFaceJob).filter(ParticipantFaceJob.face_image_id.in_(face_image_ids)).delete(synchronize_session=False)
        session.query(ParticipantFaceEmbedding).filter(ParticipantFaceEmbedding.face_image_id.in_(face_image_ids)).delete(synchronize_session=False)
    session.query(ParticipantFaceImage).filter(ParticipantFaceImage.participant_id.in_(ids)).delete(synchronize_session=False)


def acquire_admin_lock(session: Session, *, session_id: str) -> bool:
    now = utc_now()
    table = AdminSessionLock.__table__
    insert_factory = sqlite_insert if session.get_bind().dialect.name == "sqlite" else postgresql_insert
    statement = insert_factory(table).values(id=ADMIN_LOCK_ID, session_id=session_id, last_seen_at=now)
    result = session.execute(
        statement.on_conflict_do_update(
            index_elements=[table.c.id],
            set_={"session_id": session_id, "last_seen_at": now},
            where=(table.c.session_id == session_id) | (table.c.last_seen_at < now - ADMIN_SESSION_TTL),
        ).returning(table.c.session_id)
    )
    acquired_session_id = result.scalar_one_or_none()
    session.commit()
    return acquired_session_id == session_id


def force_admin_lock(session: Session, *, session_id: str) -> None:
    now = utc_now()
    table = AdminSessionLock.__table__
    insert_factory = sqlite_insert if session.get_bind().dialect.name == "sqlite" else postgresql_insert
    statement = insert_factory(table).values(id=ADMIN_LOCK_ID, session_id=session_id, last_seen_at=now)
    session.execute(
        statement.on_conflict_do_update(
            index_elements=[table.c.id],
            set_={"session_id": session_id, "last_seen_at": now},
        )
    )
    session.commit()
